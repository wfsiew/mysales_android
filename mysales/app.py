from openpyxl import load_workbook

STMT = ('insert into sales(ims_class1, year, period, cust_code, data, cust_name, clinic, ims_class2, cust_group1, cust_group2, cust_group3, '
        'corporate_group, cust_addr1, cust_addr2, cust_addr3, '
        'postal_code, area, territory, state, manager, detail_man_name, detail_man_code, '
        'zp_item_code, product_group, item_name, sales_unit, bonus_unit, sales_value) values('
        "'{0}', {1}, {2}, '{3}', '{4}', '{5}', '{6}', '{7}', '{8}', '{9}', '{10}', "
        "'{11}', '{12}', '{13}', '{14}', "
        "'{15}', '{16}', '{17}', '{18}', '{19}', '{20}', '{21}', "
        "'{22}', '{23}', '{24}', {25}, {26}, {27})"
)

def getsimilarheader(dic, key, otherkeys=[]):
    k = None
    if key in dic:
        k = dic[key]

    else:
        for s in otherkeys:
            if s in dic:
                k = dic[s]
                break

    return k

def getcellvalue(ws, i, dic, key, otherkeys=[]):
    j = getsimilarheader(dic, key, otherkeys)
    v = 0
    if j is not None:
        v = getvalue(ws.cell(row=i, column=j))

    if v is None:
        raise Exception('unable to find column {0} in row {1}'.format(key, i))

    if v == 0 and key == 'customer code':
        v = 'PHARMASERV'

    return v

def getlastrow(ws):
    i = 0
    for row in ws.iter_rows():
        k = row[0].value
        if k in [None, 'NULL']:
            break

        i += 1

    return i

def getvalue(c):
    k = c.value
    if k in [None, 'NULL']:
        k = 0

    elif isinstance(k, str) and k.find("'") >= 0:
        k = k.replace("'", "''")

    return k

def getheader(ws):
    dic = {}
    i = 0
    for row in ws.iter_rows():
        i += 1
        if i > 1: break
        j = 0
        for cell in row:
            j += 1
            k = getvalue(cell)
            if k != 0:
                dic[k.lower()] = j

    return dic

def readfile2():
    o = open('data2.sql', 'w')
    wb = load_workbook('file2.xlsx')
    ls = wb.get_sheet_names()
    for s in ls:
        procsheet(s, wb, o)

    wb.close()
    o.close()

def readfile1():
    o = open('data1.sql', 'w')
    wb = load_workbook('file1.xlsx')
    ls = wb.get_sheet_names()
    for s in ls:
        procsheet(s, wb, o)

    wb.close()
    o.close()

def procsheet(sheetname, wb, o):
    ws = wb[sheetname]
    n = getlastrow(ws)
    dic = getheader(ws)
    m = n + 1
    print('sheet {0} lastrow: {1}'.format(sheetname, n))

    x = getcellvalue
    print('start sheet {0}'.format(sheetname))

    for i in range(2, m):
        c = STMT.format(
            x(ws, i, dic, 'ims class 1'), 
            x(ws, i, dic, 'year'),
            x(ws, i, dic, 'period'),
            x(ws, i, dic, 'customer code'),
            x(ws, i, dic, 'data'),

            x(ws, i, dic, 'customer name', ['hospital name']),
            x(ws, i, dic, 'clinic'),
            x(ws, i, dic, 'ims class 2'),
            x(ws, i, dic, 'customer group 1'),
            x(ws, i, dic, 'customer group 2'),

            x(ws, i, dic, 'customer group 3'),
            x(ws, i, dic, 'corporate group'),
            x(ws, i, dic, 'customer address 1', ['address 1']),
            x(ws, i, dic, 'customer address 2', ['address 2']),
            x(ws, i, dic, 'customer address 3', ['address 3']),

            x(ws, i, dic, 'postal zip code', ['post code']),
            x(ws, i, dic, 'area'),
            x(ws, i, dic, 'territory'),
            x(ws, i, dic, 'state', 'address 4'),
            x(ws, i, dic, 'manager'),

            x(ws, i, dic, 'detailman name'),
            x(ws, i, dic, 'detailman code'),
            x(ws, i, dic, 'zp item code'),
            x(ws, i, dic, 'product group'),
            x(ws, i, dic, 'item name', ['product name']),

            x(ws, i, dic, 'sales units', ['sales qty', 'sales units sum']),
            x(ws, i, dic, 'bonus units sum'),
            x(ws, i, dic, 'sales value', ['sales value sum'])
        )

        o.write(c + '\n')
        
    print('done sheet: {0}'.format(sheetname))

def readfile():
    o = open('data.sql', 'w')
    wb = load_workbook('sample.xlsx', read_only=True)
    ws = wb['Sheet2']

    m = 8455

    i = 0
    for row in ws.rows:
        i += 1
        if i == 1: continue
        if i > m: break

        a = []
        for cell in row:
            k = cell.value
            if k is None or k == 'NULL':
                k = 0

            elif isinstance(k, str) and k.find("'") >= 0:
                k = k.replace("'", "''")

            a.append(k)

        v = ('insert into sales(ims_class1, year, period, cust_code, data, cust_name, clinic, ims_class2, cust_group1, cust_group2, cust_group3, '
            'corporate_group, cust_addr1, cust_addr2, cust_addr3, '
            'postal_code, area, territory, state, manager, detail_man_name, detail_man_code, '
            'zp_item_code, product_group, item_name, sales_unit, bonus_unit, sales_value) values('
            "'{0}', {1}, {2}, '{3}', '{4}', '{5}', '{6}', '{7}', '{8}', '{9}', '{10}', "
            "'{11}', '{12}', '{13}', '{14}', "
            "'{15}', '{16}', '{17}', '{18}', '{19}', '{20}', '{21}', "
            "'{22}', '{23}', '{24}', {25}, {26}, {27})"
        )
        c = v.format(a[0], a[1], a[2], a[3], a[4], a[5], a[6], a[7], a[8], a[9], a[10],
        a[11], a[12], a[13], a[14],
        a[15], a[16], a[17], a[18], a[19], a[20], a[21],
        a[22], a[23], a[24], a[25], a[26], a[27])
        o.write(c + '\n')

    wb.close()
    o.close()

if __name__ == '__main__':
    #readfile()
    readfile1()
    readfile2()